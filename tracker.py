from bs4 import BeautifulSoup
import requests
import re
from datetime import date
import os
import openpyxl
import enum

# These are the row labels in the excel
class labels(enum.Enum):
    Date= 1
    Title= 2
    URL= 3
    Price= 4
    Sale= 5

# This function takes in the URL that we want to webscrape
# The sheet and row index allows us modify the corresponding information about the game
def extract(URL, sheet, row):
    HEADERS = ({'User-Agent':
                'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/44.0.2403.157 Safari/537.36',
                                'Accept-Language': 'en-US, en;q=0.5'})
    webpage= requests.get(URL, headers= HEADERS, cookies= cookies)
    soup= BeautifulSoup(webpage.content, "lxml")

    # This will set the date at which information was updated
    sheet.cell(row= row, column= labels.Date.value).value = today.strftime("%B %d")

    # Trying to get the title of the game

    try:
        title= soup.find("div", attrs={"id" : 'appHubAppName', "class" : "apphub_AppName"})
        title= title.string.strip()
    except AttributeError:
        title = "NA"

    sheet.cell(row= row, column= labels.Title.value).value = title

    # This is to find the price of the item
    try:
        parent= soup.find("div", attrs= {"class" : "game_area_purchase_game_wrapper"})

    except AttributeError:
        parent= 0
        orig_price= "NA"
        disc_price= "NA"

    if (parent != 0):
        try:
            orig_price= parent.find("div", attrs={"class" : "discount_original_price"}, recursive= True)
            disc_price = parent.find("div", attrs={"class" : "discount_final_price"}, recursive= True)
            orig_price= orig_price.string.strip()
            disc_price= disc_price.string.strip()
        except AttributeError:
            try:
                orig_price= parent.find("div", attrs={"class" : 'game_purchase_price price',
                "data-price-final" : re.compile("\d*")}, recursive= True).string.strip()
                disc_price= "NA"
            except:
                orig_price= "NA"
                disc_price= "NA"



    sheet.cell(row= row, column = labels.Price.value).value = orig_price

    sheet.cell(row= row, column = labels.Sale.value).value = disc_price



if __name__ == "__main__":
    today= date.today()
    cookies = {'birthtime': '568022401'}

    cur_path= os.path.dirname(__file__)

    loc_wb= os.path.join(cur_path, "games.xlsx")
    try:
        wb = openpyxl.load_workbook(loc_wb)
    except FileNotFoundError:
        wb = openpyxl.Workbook()

    sheet = wb.active


    for label in (labels):
        cur_cell= sheet.cell(row= 1, column = label.value)
        if (cur_cell.value != "None"):
            cur_cell.value = label.name


    for i in range(2, sheet.max_row+1):
        cur_cell = sheet.cell(row = i, column = labels.URL.value)
        if (not (cur_cell.value == "None" or cur_cell.value == None)):
            extract(cur_cell.value, sheet, i)

    sheet.column_dimensions["B"].width= 50




    wb.save(loc_wb)
